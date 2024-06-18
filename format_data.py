from openpyxl import *
from openpyxl.styles import *
import pandas as pd
from datetime import date

def format(raw_dpp_filepath):

    # Read from the raw DPP Excel file
    df = pd.read_excel(raw_dpp_filepath, sheet_name='Sheet0')

    # Rename Column Names (Reflects Leah's Excel Sheet)
    df.rename(columns={
        "Service Address Line 1": "Service Address",
        "Service Postal Code": "Postal Code",
        "State": "Province",
        "Service Call Date": "Schedule Date",
        "Dispatch Status": "Status"
    }, inplace=True)

    # Delete Redundant Columns
    df.drop(columns=[
    "Service Provider", "Logistics Provider", 
    "LOB", "End Service Window", 
    "DPS Type", "Call Type", "Service Level",
    "Customer Number", "Customer Secondary Contact Name", 
    "Customer Secondary Contact Phone Number", 
    "Customer Secondary Contact Email Address", 
    "Service Address Line 2", "Service Address Line 3",
    "Service Address Line 4", "Part Number", "Part Quantity", 
    "Part Status", "Part Status Date", "Carrier Name",
    "Waybill Number", "Closure Date", "Warranty Invoice Date", 
    "Warranty Invoice Number", "Original Order BUID", 
    "Reply Code", "Service Request Type", "Product Classification", 
    "Report Description", "Service Type"
    ], inplace=True)


    # Add New Columns for Manual Entry
    df["Zone Uplift"] = ''
    df["Overnight"] = ''
    df["Shift Uplift"] = ''
    df["PM Confirmation"] = ''
    df["Completed Date"] = ''

    # Restructure Column Order
    new_column_order = [
        "Dispatch Number", "Status",
        "Service Address", "City", "Postal Code",
        "Province", "Country", 
        "Project Number", "Product Name",
        "Product Model", "Service Tag",
        "Service SKU", "Corrected SKU",
        "Comments to Vendor", "Zone Uplift",
        "Overnight", "Shift Uplift",
        "Engineer Assigned", "Engineer Id",
        "PM Contact", "PM Phone",
        "PM Email", "Customer Name",
        "Customer Contact Phone Number", "Customer Contact Email Address",
        "Schedule Date", "PM Confirmation",
        "Status", "Completed Date"
    ]    
    
   # current_date = date.today()


    # Reorder the DataFrame columns
    df = df.reindex(columns=new_column_order) 


    # Remove Quantity from Service SKU
    text_to_remove = "Qty:1"
    df["Service SKU"] = df["Service SKU"].str.replace(text_to_remove, "")

    # Save the DataFrame to an Excel file
    filename = f"workorders_{date.today()}.xlsx"
    df.to_excel(filename, sheet_name='Sheet0', index=False)

    # Load the formatted workbook
    wb = load_workbook(filename)
    ws = wb['Sheet0']

    # Style and format cells
    style_cells(ws)
    format_cells(ws)

    # Save the formatted workbook
    wb.save(filename)

    # Return filename
    return filename


"""
Styles cells
    - Green fill to indicate new entries in the master sheet (when merged)
    - Red fill to indicate missing values (except manual entry columns)
"""
def style_cells(ws):
    # Columns to ignore if they're empty
    ignore_cols = [
        "Missing Information", "Zone Uplift", 
        "Overnight", "Shift Uplift", 
        "PM Confirmation", "Completed Date"
    ]

    # Column to indicate missing value(s) in a row
    # Make a dictionary to map column names to indices
    col_indices = {
        cell.value: cell.column for cell in ws[1]
    }
    isMissing = col_indices.get("Missing Information")

    # Apply a green fill for all new entries
    for col in ws.columns:
        for cell in col:
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            if cell not in ignore_cols:
                if cell.value is None:
                    # If a cell is empty, apply a red fill and set isMissing to true
                    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCB", fill_type="solid")
                    ws.cell(row=cell.row, column=isMissing, value="True")


"""
Formats cells
    - Applies text wrap and auto-fits column widths
"""
def format_cells(ws):
    # Apply text wrap and auto-fit column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            cell.alignment = Alignment(wrapText=True)
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.1  # Adjust column width
        ws.column_dimensions[col[0].column_letter].width = adjusted_width